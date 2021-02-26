from hasher import RecursiveDirectoriesHasher
from handlers import rdh_output_handler
import functools


def get_xlsx_content(file):
    from openpyxl import load_workbook
    wb = load_workbook(file)
    ws = wb.active
    data = []
    for row in ws.iter_rows():
        data_row = []
        for cell in row:
            data_row.append(cell.value)
        data.append(data_row)
    return data[1:]


def verify(directory, export):
    rdh = RecursiveDirectoriesHasher(directory)
    rdh_data = rdh.start_no_cb()
    rdh_data = rdh_output_handler(rdh_data, xlsx_mode=False)

    xlsx_data = get_xlsx_content(export)
    return rdh_data == xlsx_data
